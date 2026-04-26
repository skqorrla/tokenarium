"""xlsx 시드 적재.

`docs/<table>.xlsx` 파일을 읽어 동명 테이블에 INSERT OR IGNORE 로 적재.
파일명이 곧 테이블명이며, xlsx 첫 행이 컬럼명. 신규 시드 테이블 추가는
xlsx 한 장을 docs/ 에 떨어뜨리면 끝.

전략: INSERT OR IGNORE — 첫 적재만 반영, 재실행 시 PK/UNIQUE 충돌은 조용히 무시.
"""

from __future__ import annotations

import sqlite3
from pathlib import Path

from openpyxl import load_workbook


def seed_from_xlsx(db_path: str, docs_dir: Path) -> dict[str, int]:
    """docs_dir 의 모든 *.xlsx → 동명 테이블에 INSERT OR IGNORE.

    Returns: {table_name: 적재_시도_행수} (실제 INSERT 행수가 아니라 xlsx 의 데이터 행수.
    OR IGNORE 로 충돌은 조용히 스킵되므로 시도/성공 구분이 의미 없음.)
    """
    docs_dir = Path(docs_dir)
    if not docs_dir.is_dir():
        raise FileNotFoundError(f"docs dir not found: {docs_dir}")

    counts: dict[str, int] = {}
    conn = sqlite3.connect(db_path)
    try:
        existing = {
            row[0]
            for row in conn.execute(
                "SELECT name FROM sqlite_master WHERE type='table'"
            )
        }
        for xlsx_path in sorted(docs_dir.glob("*.xlsx")):
            table = xlsx_path.stem
            if table not in existing:
                print(f"[seed] skip {xlsx_path.name} — table '{table}' 없음")
                continue
            counts[table] = _load_xlsx_to_table(conn, xlsx_path, table)
        conn.commit()
    finally:
        conn.close()
    return counts


def _load_xlsx_to_table(conn: sqlite3.Connection, xlsx_path: Path, table: str) -> int:
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = next(rows_iter)
    except StopIteration:
        return 0

    headers = [h for h in headers if h is not None]
    cols = ", ".join(headers)
    placeholders = ", ".join("?" * len(headers))
    sql = f"INSERT OR IGNORE INTO {table} ({cols}) VALUES ({placeholders})"

    data = [
        tuple(row[: len(headers)])
        for row in rows_iter
        if row and any(c is not None for c in row[: len(headers)])
    ]
    if data:
        conn.executemany(sql, data)
    return len(data)
